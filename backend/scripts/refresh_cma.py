"""Build the CMA snapshot from the licensed J.P. Morgan LTCMA CHF matrix (Phase 2).

DEV/SEED ONLY — off the request path. Reads the J.P. Morgan 2026 Long-Term
Capital Market Assumptions **CHF** matrix (a licensed file kept locally in
``docs/references/``, git-ignored — DATA.md), extracts the ten asset classes our
universe uses, and writes the validated, positive-semidefinite
``data/snapshots/cma.json``. Every figure traces to that file; none is invented.

The optimiser consumes ``expected_return`` (JPM's *arithmetic* 2026 return) as its
``mu``; projections use ``compound_return`` (JPM's *compound* 2026 return). A
row-label guard fails loudly if a future JPM edition reorders its rows, so the
fixed row indices below can never silently read the wrong asset.

Run from ``backend/``::

    python -m scripts.refresh_cma
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import numpy as np
import openpyxl

from app.snapshots.schemas import CMA, AssetClass, AssetClassAssumption, Citation
from app.snapshots.validate import assert_cma_psd

_BACKEND = Path(__file__).resolve().parents[1]
_REPO = _BACKEND.parent
XLSX_PATH = _REPO / "docs" / "references" / "jpm-ltcma-2026-chf.xlsx"
OUT_PATH = _BACKEND / "data" / "snapshots" / "cma.json"

SHEET = "CHF"
# Sheet columns (1-indexed). The header is a diagonal "staircase" of merged cells
# — B8:C8 "Compound Return 2026", B7:D7 "Arithmetic Return 2026", B6:E6 "Annualized
# Volatility", B5:F5 "Compound Return 2025" — where each label's RIGHTMOST column is
# its data column. So: C=Compound 2026, D=Arithmetic 2026, E=Volatility, F=Compound
# 2025 (the PRIOR edition; intentionally unused — reading it for the compound was a
# bug that let cash show arithmetic < compound).
_COL_LABEL = 2
_COL_COMPOUND = 3  # C — current-edition compound return (projections)
_COL_ARITHMETIC = 4  # D — arithmetic 2026 (the optimiser's mu)
_COL_VOLATILITY = 5  # E — annualized volatility
# The correlation block starts at column 7 in the same row order as the assets;
# the column for the asset at sheet row r is therefore r - 2.
def _corr_col(sheet_row: int) -> int:
    return sheet_row - 2


# our asset class -> (JPM sheet row, our display label, the JPM row label we
# expect to find there [guard], mapping note for the citation)
_MAPPING: tuple[tuple[AssetClass, int, str, str, str], ...] = (
    (AssetClass.CASH_CHF, 10, "CHF cash", "Swiss Cash", "JPM 'Swiss Cash'"),
    (AssetClass.SWISS_BONDS, 11, "Swiss government bonds", "Swiss Government Bonds",
     "JPM 'Swiss Government Bonds'"),
    (AssetClass.GLOBAL_AGG_BONDS, 22, "Global bonds (CHF-hedged)", "World Government Bonds hedged",
     "JPM 'World Government Bonds hedged' — closest single global CHF-hedged bond series "
     "(proxy for the global aggregate; govt-only, so a touch conservative on yield)"),
    (AssetClass.SWISS_EQUITY, 29, "Swiss equity", "Swiss Equity", "JPM 'Swiss Equity'"),
    (AssetClass.US_EQUITY, 30, "US large-cap equity", "U.S. Large Cap", "JPM 'U.S. Large Cap'"),
    (AssetClass.EUROPE_EQUITY, 33, "European equity", "European Large Cap",
     "JPM 'European Large Cap'"),
    (AssetClass.JAPAN_EQUITY, 39, "Japanese equity", "Japanese Equity", "JPM 'Japanese Equity'"),
    (AssetClass.EM_EQUITY, 42, "Emerging-market equity", "Emerging Markets Equity",
     "JPM 'Emerging Markets Equity'"),
    (AssetClass.WORLD_EQUITY, 43, "Global (all-country) equity", "AC World Equity",
     "JPM 'AC World Equity'"),
    (AssetClass.GOLD, 56, "Gold", "Gold", "JPM 'Gold'"),
)

_SOURCE_AS_OF = date(2025, 10, 20)  # JPM 2026 LTCMA publication date
_SOURCE_URL = (
    "https://am.jpmorgan.com/ch/en/asset-management/institutional/insights/"
    "portfolio-insights/ltcma/"
)
_HORIZON_YEARS = 13  # JPM states a 10-15 year horizon


def build_cma() -> CMA:
    """Extract the ten-class CMA from the JPM CHF matrix and return it validated."""
    if not XLSX_PATH.exists():
        raise SystemExit(
            f"JPM CHF matrix not found at {XLSX_PATH}.\n"
            "Download the CHF 'matrices by currency' Excel from JPM Asset Management's "
            "Long-Term Capital Market Assumptions page and save it there (it is git-ignored)."
        )
    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True)[SHEET]

    n = len(_MAPPING)
    assumptions: list[AssetClassAssumption] = []
    corr = np.full((n, n), np.nan)

    for i, (asset_class, row, label, expected_label, _note) in enumerate(_MAPPING):
        actual_label = str(ws.cell(row=row, column=_COL_LABEL).value or "")
        if expected_label.lower() not in actual_label.lower():
            raise SystemExit(
                f"row-label guard: expected '{expected_label}' at sheet row {row}, "
                f"found '{actual_label}'. The JPM layout changed — fix _MAPPING."
            )
        arithmetic = ws.cell(row=row, column=_COL_ARITHMETIC).value
        volatility = ws.cell(row=row, column=_COL_VOLATILITY).value
        compound = ws.cell(row=row, column=_COL_COMPOUND).value
        if arithmetic is None or volatility is None or compound is None:
            raise SystemExit(f"missing return/vol for '{label}' at sheet row {row}")
        assumptions.append(
            AssetClassAssumption(
                asset_class=asset_class,
                label=label,
                expected_return=round(float(arithmetic) / 100.0, 6),
                compound_return=round(float(compound) / 100.0, 6),
                volatility=round(float(volatility) / 100.0, 6),
            )
        )
        for j, (_ac, row_j, *_rest) in enumerate(_MAPPING):
            value = ws.cell(row=row, column=_corr_col(row_j)).value
            if value is not None:
                corr[i, j] = float(value)

    # JPM stores only the lower triangle — mirror it into the upper triangle.
    for i in range(n):
        for j in range(i + 1, n):
            if np.isnan(corr[i, j]):
                corr[i, j] = corr[j, i]
    if np.isnan(corr).any():
        raise SystemExit("correlation matrix has gaps after mirroring")

    # Round, then re-impose exact symmetry and a unit diagonal (rounding can nudge
    # an entry by ~1e-6, which the strict schema validator would reject).
    rounded = [[round(float(corr[i, j]), 6) for j in range(n)] for i in range(n)]
    for i in range(n):
        rounded[i][i] = 1.0
        for j in range(i + 1, n):
            rounded[j][i] = rounded[i][j]

    citation = Citation(
        source="J.P. Morgan Asset Management — 2026 Long-Term Capital Market Assumptions (CHF)",
        url=_SOURCE_URL,
        as_of=_SOURCE_AS_OF,
        note=(
            "Licensed CHF matrix kept locally in docs/references/ (git-ignored). "
            "expected_return = JPM arithmetic 2026 return; compound_return = JPM compound 2026 "
            "return; volatility = JPM annualized volatility. Class->JPM-row mapping: "
            + "; ".join(f"{ac.value}={note}" for ac, _r, _l, _e, note in _MAPPING)
            + "."
        ),
    )

    return CMA(
        schema_version=1,
        source=citation,
        base_currency="CHF",
        horizon_years=_HORIZON_YEARS,
        assumptions=tuple(assumptions),
        correlation=tuple(tuple(r) for r in rounded),
    )


def main() -> None:
    cma = build_cma()
    assert_cma_psd(cma)  # never ship a risk model the optimiser could exploit
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(cma.model_dump_json(indent=2) + "\n", encoding="utf-8")
    print(
        f"wrote {OUT_PATH.relative_to(_REPO)} — {len(cma.assumptions)} asset classes, "
        f"base {cma.base_currency}, horizon {cma.horizon_years}y, PSD ✓"
    )


if __name__ == "__main__":
    main()
